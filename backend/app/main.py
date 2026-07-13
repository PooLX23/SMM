import json
from datetime import datetime
from io import BytesIO

from fastapi import FastAPI, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select, or_, func
from sqlalchemy.orm import Session, selectinload
from sqlalchemy.exc import IntegrityError

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .deps import get_db, get_current_user, CurrentUser, require_reception
from .models import (
    Shipment, ShipmentStatus, Direction, ShipmentEvent, Carrier, CostCenter,
    IncomingShipment, AddressBookEntry
)
from .schemas import (
    ShipmentCreate, ShipmentOut, ShipmentShip, CostCenterOut, SimpleDictItem,
    IncomingShipmentCreate, IncomingShipmentOut, AddressBookEntryCreate,
    AddressBookEntryUpdate, AddressBookEntryOut
)
from .internal_no import next_outgoing_internal_no, next_incoming_internal_no

app = FastAPI(title="Courier Registry API", version="0.1.0")


class IncomingRecipientChange(BaseModel):
    recipient_upn: str
    recipient_name: str


class ShipmentShippingChange(BaseModel):
    carrier_id: str
    carrier_tracking_no: str


def add_event(db: Session, shipment_id, event_type: str, user_upn: str, payload: dict | None = None):
    ev = ShipmentEvent(
        shipment_id=shipment_id,
        event_type=event_type,
        created_by_upn=user_upn,
        payload_json=json.dumps(payload or {}, ensure_ascii=False),
    )
    db.add(ev)


def _shipment_event_times(db: Session, shipment_id) -> dict:
    rows = (
        db.execute(
            select(ShipmentEvent.event_type, func.max(ShipmentEvent.created_at))
            .where(ShipmentEvent.shipment_id == shipment_id)
            .where(
                ShipmentEvent.event_type.in_(
                    [
                        "CANCELLED",
                        "CANCELLED_AFTER_SHIPPED",
                        "SHIPPING_CHANGED",
                    ]
                )
            )
            .group_by(ShipmentEvent.event_type)
        )
        .all()
    )

    by_type = {event_type: event_at for event_type, event_at in rows}
    return {
        "cancelled_at": by_type.get("CANCELLED"),
        "cancelled_after_shipped_at": by_type.get("CANCELLED_AFTER_SHIPPED"),
        "shipping_changed_at": by_type.get("SHIPPING_CHANGED"),
    }


def shipment_to_out(sh: Shipment, db: Session | None = None) -> dict:
    """
    Jawne mapowanie ORM -> dict zgodny z ShipmentOut.
    Dzięki temu dokładamy:
      - cost_center_code
      - cost_center_name
      - czasy statusów z historii ShipmentEvent
    """
    cc = sh.cost_center
    event_times = _shipment_event_times(db, sh.id) if db is not None else {}
    return {
        "id": str(sh.id),
        "internal_no": sh.internal_no,
        "direction": sh.direction,
        "status": sh.status,

        "recipient_name": sh.recipient_name,
        "recipient_email": sh.recipient_email,
        "recipient_phone": sh.recipient_phone,
        "recipient_postal_code": sh.recipient_postal_code,
        "recipient_city": sh.recipient_city,
        "recipient_country": sh.recipient_country,
        "recipient_street": sh.recipient_street,

        "contents": sh.contents,
        "vin": sh.vin,
        "plate_no": sh.plate_no,

        "requested_by_upn": sh.requested_by_upn,
        "requested_by_name": sh.requested_by_name,

        "cost_center_id": str(sh.cost_center_id),
        "cost_center_code": (cc.code if cc else None),
        "cost_center_name": (cc.name if cc else None),

        "carrier_id": str(sh.carrier_id) if sh.carrier_id else None,
        "carrier_tracking_no": sh.carrier_tracking_no,

        "received_at": sh.received_at,
        "shipped_at": sh.shipped_at,
        "cancelled_at": event_times.get("cancelled_at"),
        "cancelled_after_shipped_at": event_times.get("cancelled_after_shipped_at"),
        "shipping_changed_at": event_times.get("shipping_changed_at"),

        "created_at": sh.created_at,
        "updated_at": sh.updated_at,
    }


def incoming_to_out(x: IncomingShipment) -> dict:
    carrier_name = ""
    try:
        carrier = getattr(x, "carrier", None)
        carrier_name = getattr(carrier, "name", "") if carrier else ""
    except Exception:
        carrier_name = ""

    return {
        "id": str(x.id),
        "internal_no": x.internal_no,
        "direction": x.direction,
        "status": x.status,

        "carrier_id": str(x.carrier_id) if x.carrier_id else None,
        "carrier_name": carrier_name or None,
        "carrier_tracking_no": x.carrier_tracking_no,

        "sender_name": x.sender_name,
        "contents": x.contents,

        "recipient_upn": x.recipient_upn,
        "recipient_name": x.recipient_name,

        "received_at": x.received_at,
        "picked_up_at": x.picked_up_at,

        "created_at": x.created_at,
        "updated_at": x.updated_at,
    }


# ======================================================
# XLSX EXPORT (recepcja)
# ======================================================

def excel_safe(val):
    # Excel/openpyxl nie obsługuje timezone-aware datetimes/times
    try:
        if isinstance(val, datetime) and getattr(val, "tzinfo", None) is not None:
            return val.replace(tzinfo=None)
    except Exception:
        pass
    return val


def build_xlsx(sheet_name: str, headers: list[str], rows: list[list]):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.freeze_panes = "A2"

    header_font = Font(bold=True)
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=excel_safe(val))

    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 60)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


@app.get("/api/reception/export/outgoing.xlsx")
def export_outgoing_xlsx(
    status: ShipmentStatus | None = Query(default=None),
    q: str | None = Query(default=None),
    limit: int = Query(default=500, ge=1, le=5000),
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(require_reception),
):
    # 1:1 z /api/shipments (filtry q/status + sort), tylko wymuszamy OUTGOING
    stmt = (
        select(Shipment)
        .options(selectinload(Shipment.cost_center), selectinload(Shipment.carrier))
        .where(Shipment.direction == Direction.OUTGOING)
        .order_by(Shipment.created_at.desc())
        .limit(limit)
    )

    if status:
        stmt = stmt.where(Shipment.status == status)

    if q:
        like = f"%{q.strip()}%"
        stmt = stmt.where(
            or_(
                Shipment.internal_no.ilike(like),
                Shipment.recipient_name.ilike(like),
                Shipment.recipient_email.ilike(like),
                Shipment.recipient_phone.ilike(like),
                Shipment.carrier_tracking_no.ilike(like),

                Shipment.recipient_street.ilike(like),
                Shipment.recipient_city.ilike(like),
                Shipment.recipient_postal_code.ilike(like),
            )
        )

    rows_db = db.execute(stmt).scalars().all()

    headers = [
        "Internal No", "Status",
        "Recipient name", "Recipient email", "Recipient phone",
        "Recipient street", "Recipient postal code", "Recipient city", "Recipient country",
        "Contents", "VIN", "Plate No",
        "Requested by (UPN)", "Requested by (Name)",
        "Cost center code", "Cost center name",
        "Carrier", "Carrier tracking no",
        "Received at", "Shipped at",
        "Created at", "Updated at",
    ]

    rows = []
    for sh in rows_db:
        cc = sh.cost_center
        carrier = sh.carrier
        rows.append([
            sh.internal_no, str(sh.status),
            sh.recipient_name, sh.recipient_email, sh.recipient_phone,
            sh.recipient_street, sh.recipient_postal_code, sh.recipient_city, sh.recipient_country,
            sh.contents, sh.vin, sh.plate_no,
            sh.requested_by_upn, sh.requested_by_name,
            (cc.code if cc else None), (cc.name if cc else None),
            (carrier.name if carrier else None), sh.carrier_tracking_no,
            sh.received_at, sh.shipped_at,
            sh.created_at, sh.updated_at,
        ])

    bio = build_xlsx("Wyslane", headers, rows)
    filename = f"wyslane_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/reception/export/incoming.xlsx")
def export_incoming_xlsx(
    status: ShipmentStatus | None = Query(default=None),
    q: str | None = Query(default=None),
    limit: int = Query(default=500, ge=1, le=5000),
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(require_reception),
):
    # 1:1 z /api/incoming-shipments (filtry q/status + sort)
    stmt = (
        select(IncomingShipment)
        .options(selectinload(IncomingShipment.carrier))
        .order_by(func.coalesce(IncomingShipment.received_at, IncomingShipment.created_at).desc())
        .limit(limit)
    )

    if status:
        stmt = stmt.where(IncomingShipment.status == status)

    if q:
        like = f"%{q.strip()}%"
        stmt = stmt.where(
            or_(
                IncomingShipment.internal_no.ilike(like),
                IncomingShipment.carrier_tracking_no.ilike(like),
                IncomingShipment.sender_name.ilike(like),
                IncomingShipment.recipient_name.ilike(like),
                IncomingShipment.recipient_upn.ilike(like),
                IncomingShipment.contents.ilike(like),
            )
        )

    rows_db = db.execute(stmt).scalars().all()

    headers = [
        "Internal No", "Status",
        "Carrier", "Carrier tracking no",
        "Sender name", "Contents",
        "Recipient UPN", "Recipient name",
        "Received at", "Picked up at",
        "Created at", "Updated at",
    ]

    rows = []
    for x in rows_db:
        carrier = getattr(x, "carrier", None)
        rows.append([
            x.internal_no, str(x.status),
            (carrier.name if carrier else None), x.carrier_tracking_no,
            x.sender_name, x.contents,
            x.recipient_upn, x.recipient_name,
            x.received_at, x.picked_up_at,
            x.created_at, x.updated_at,
        ])

    bio = build_xlsx("Odebrane", headers, rows)
    filename = f"odebrane_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/health")
def health():
    return {"ok": True}


@app.get("/api/me")
def me(user: CurrentUser = Depends(get_current_user)):
    return {"upn": user.upn, "name": user.name, "is_reception": user.is_reception, "groups_count": len(user.groups)}


@app.get("/api/cost-centers", response_model=list[CostCenterOut])
def list_cost_centers(db: Session = Depends(get_db), user: CurrentUser = Depends(get_current_user)):
    rows = db.execute(select(CostCenter).where(CostCenter.active == True).order_by(CostCenter.code)).scalars().all()
    return rows


@app.get("/api/carriers", response_model=list[SimpleDictItem])
def list_carriers(db: Session = Depends(get_db), user: CurrentUser = Depends(get_current_user)):
    rows = db.execute(select(Carrier).where(Carrier.active == True).order_by(Carrier.name)).scalars().all()
    return rows


@app.get("/api/address-book", response_model=list[AddressBookEntryOut])
def list_address_book(
    q: str | None = Query(default=None),
    limit: int = Query(default=200, ge=1, le=500),
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(get_current_user),
):
    stmt = select(AddressBookEntry).order_by(AddressBookEntry.recipient_name.asc()).limit(limit)

    if q:
        like = f"%{q.strip()}%"
        stmt = stmt.where(
            or_(
                AddressBookEntry.recipient_name.ilike(like),
                AddressBookEntry.recipient_email.ilike(like),
                AddressBookEntry.recipient_phone.ilike(like),
                AddressBookEntry.recipient_street.ilike(like),
                AddressBookEntry.recipient_city.ilike(like),
                AddressBookEntry.recipient_postal_code.ilike(like),
            )
        )

    return db.execute(stmt).scalars().all()


@app.post("/api/address-book", response_model=AddressBookEntryOut)
def create_address_book_entry(
    body: AddressBookEntryCreate,
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(require_reception),
):
    now = datetime.utcnow()
    entry = AddressBookEntry(
        recipient_name=body.recipient_name.strip(),
        recipient_email=str(body.recipient_email).strip(),
        recipient_phone=body.recipient_phone.strip(),
        recipient_street=body.recipient_street.strip(),
        recipient_country=body.recipient_country.strip().upper(),
        recipient_postal_code=body.recipient_postal_code.strip(),
        recipient_city=body.recipient_city.strip(),
        created_at=now,
        updated_at=now,
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return entry


@app.put("/api/address-book/{entry_id}", response_model=AddressBookEntryOut)
def update_address_book_entry(
    entry_id: str,
    body: AddressBookEntryUpdate,
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(require_reception),
):
    entry = db.get(AddressBookEntry, entry_id)
    if not entry:
        raise HTTPException(404, "Address book entry not found")

    entry.recipient_name = body.recipient_name.strip()
    entry.recipient_email = str(body.recipient_email).strip()
    entry.recipient_phone = body.recipient_phone.strip()
    entry.recipient_street = body.recipient_street.strip()
    entry.recipient_country = body.recipient_country.strip().upper()
    entry.recipient_postal_code = body.recipient_postal_code.strip()
    entry.recipient_city = body.recipient_city.strip()
    entry.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(entry)
    return entry


@app.delete("/api/address-book/{entry_id}")
def delete_address_book_entry(
    entry_id: str,
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(require_reception),
):
    entry = db.get(AddressBookEntry, entry_id)
    if not entry:
        raise HTTPException(404, "Address book entry not found")

    db.delete(entry)
    db.commit()
    return {"status": "OK"}


@app.post("/api/shipments", response_model=ShipmentOut)
def create_shipment(
    body: ShipmentCreate,
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(get_current_user),
):
    try:
        internal_no = next_outgoing_internal_no(db)

        sh = Shipment(
            internal_no=internal_no,
            direction=Direction.OUTGOING,
            status=ShipmentStatus.CREATED,

            recipient_name=body.recipient_name,
            recipient_email=str(body.recipient_email),
            recipient_phone=body.recipient_phone,
            recipient_postal_code=body.recipient_postal_code,
            recipient_city=body.recipient_city,
            recipient_country=body.recipient_country.upper(),
            recipient_street=body.recipient_street,

            contents=body.contents,
            vin=body.vin,
            plate_no=body.plate_no,

            requested_by_upn=user.upn,
            requested_by_name=user.name,

            cost_center_id=body.cost_center_id,
            updated_at=datetime.utcnow(),
        )

        db.add(sh)
        db.flush()  # potrzebne żeby mieć sh.id zanim dodamy event

        add_event(db, sh.id, "CREATED", user.upn, {"internal_no": internal_no})

        db.commit()
        db.refresh(sh)

        # doładuj cost center, żeby od razu zwrócić code/name
        sh = (
            db.execute(
                select(Shipment)
                .options(selectinload(Shipment.cost_center))
                .where(Shipment.id == sh.id)
            )
            .scalar_one()
        )
        return shipment_to_out(sh, db)

    except IntegrityError:
        db.rollback()
        raise HTTPException(409, "Duplicate internal number (retry)")
    except Exception:
        db.rollback()
        raise


@app.get("/api/shipments/by-internal/{internal_no}", response_model=ShipmentOut)
def get_by_internal(internal_no: str, db: Session = Depends(get_db), user: CurrentUser = Depends(get_current_user)):
    stmt = (
        select(Shipment)
        .options(selectinload(Shipment.cost_center))
        .where(Shipment.internal_no == internal_no)
    )
    sh = db.execute(stmt).scalar_one_or_none()
    if not sh:
        raise HTTPException(404, "Shipment not found")
    return shipment_to_out(sh, db)


@app.get("/api/shipments", response_model=list[ShipmentOut])
def search_shipments(
    status: ShipmentStatus | None = Query(default=None),
    q: str | None = Query(default=None),
    limit: int = Query(default=200, ge=1, le=500),
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(get_current_user),
):
    stmt = (
        select(Shipment)
        .options(selectinload(Shipment.cost_center))
        .order_by(Shipment.created_at.desc())
        .limit(limit)
    )

    if status:
        stmt = stmt.where(Shipment.status == status)

    if q:
        like = f"%{q.strip()}%"
        stmt = stmt.where(
            or_(
                Shipment.internal_no.ilike(like),
                Shipment.recipient_name.ilike(like),
                Shipment.recipient_email.ilike(like),
                Shipment.recipient_phone.ilike(like),
                Shipment.carrier_tracking_no.ilike(like),

                Shipment.recipient_street.ilike(like),
                Shipment.recipient_city.ilike(like),
                Shipment.recipient_postal_code.ilike(like),
            )
        )

    rows = db.execute(stmt).scalars().all()
    return [shipment_to_out(sh, db) for sh in rows]


@app.get("/api/my-shipments", response_model=list[ShipmentOut])
def my_shipments(
    status: ShipmentStatus | None = Query(default=None),
    q: str | None = Query(default=None),
    limit: int = Query(default=100, ge=1, le=300),
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(get_current_user),
):
    stmt = (
        select(Shipment)
        .options(selectinload(Shipment.cost_center))
        .where(Shipment.requested_by_upn == user.upn)
        .order_by(Shipment.created_at.desc())
        .limit(limit)
    )

    if status:
        stmt = stmt.where(Shipment.status == status)

    if q:
        like = f"%{q.strip()}%"
        stmt = stmt.where(
            or_(
                Shipment.internal_no.ilike(like),
                Shipment.recipient_name.ilike(like),
                Shipment.recipient_email.ilike(like),
                Shipment.recipient_phone.ilike(like),
                Shipment.carrier_tracking_no.ilike(like),

                Shipment.recipient_street.ilike(like),
                Shipment.recipient_city.ilike(like),
                Shipment.recipient_postal_code.ilike(like),
            )
        )

    rows = db.execute(stmt).scalars().all()
    return [shipment_to_out(sh, db) for sh in rows]


@app.post("/api/shipments/{shipment_id}/receive", response_model=ShipmentOut)
def receive_at_reception(
    shipment_id: str,
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(require_reception),
):
    sh = db.get(Shipment, shipment_id)
    if not sh:
        raise HTTPException(404, "Shipment not found")

    if sh.status == ShipmentStatus.AT_RECEPTION:
        sh = (
            db.execute(
                select(Shipment)
                .options(selectinload(Shipment.cost_center))
                .where(Shipment.id == sh.id)
            )
            .scalar_one()
        )
        return shipment_to_out(sh, db)

    if sh.status in (
        ShipmentStatus.SHIPPED,
        ShipmentStatus.SHIPPING_CHANGED,
        ShipmentStatus.CANCELLED,
        ShipmentStatus.CANCELLED_AFTER_SHIPPED,
    ):
        raise HTTPException(400, f"Cannot receive when status={sh.status}")

    try:
        sh.status = ShipmentStatus.AT_RECEPTION
        sh.received_at = datetime.utcnow()
        sh.updated_at = datetime.utcnow()
        add_event(db, sh.id, "AT_RECEPTION", user.upn)

        db.commit()
        db.refresh(sh)

        # mail (doładuj relacje)
        try:
            from .notify import notify_status_change
            sh_mail = (
                db.execute(
                    select(Shipment)
                    .options(selectinload(Shipment.carrier), selectinload(Shipment.cost_center))
                    .where(Shipment.id == sh.id)
                )
                .scalar_one()
            )
            notify_status_change(sh_mail, "AT_RECEPTION")
        except Exception:
            pass

        sh = (
            db.execute(
                select(Shipment)
                .options(selectinload(Shipment.cost_center))
                .where(Shipment.id == sh.id)
            )
            .scalar_one()
        )
        return shipment_to_out(sh, db)

    except Exception:
        db.rollback()
        raise


@app.post("/api/shipments/{shipment_id}/ship", response_model=ShipmentOut)
def ship(
    shipment_id: str,
    body: ShipmentShip,
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(require_reception),
):
    sh = db.get(Shipment, shipment_id)
    if not sh:
        raise HTTPException(404, "Shipment not found")

    if sh.status in (ShipmentStatus.SHIPPED, ShipmentStatus.SHIPPING_CHANGED):
        raise HTTPException(400, "Shipment is already shipped. Use change-shipping endpoint.")

    if sh.status in (ShipmentStatus.CANCELLED, ShipmentStatus.CANCELLED_AFTER_SHIPPED):
        raise HTTPException(400, "Cannot ship cancelled shipment")

    if not body.carrier_tracking_no.strip():
        raise HTTPException(400, "carrier_tracking_no is required")

    try:
        sh.carrier_id = body.carrier_id
        sh.carrier_tracking_no = body.carrier_tracking_no.strip()
        sh.status = ShipmentStatus.SHIPPED
        sh.shipped_at = datetime.utcnow()
        sh.updated_at = datetime.utcnow()

        add_event(
            db,
            sh.id,
            "SHIPPED",
            user.upn,
            {
                "carrier_id": str(body.carrier_id),
                "carrier_tracking_no": sh.carrier_tracking_no,
            },
        )

        db.commit()
        db.refresh(sh)

        # mail (doładuj relacje)
        try:
            from .notify import notify_status_change
            sh_mail = (
                db.execute(
                    select(Shipment)
                    .options(selectinload(Shipment.carrier), selectinload(Shipment.cost_center))
                    .where(Shipment.id == sh.id)
                )
                .scalar_one()
            )
            notify_status_change(sh_mail, "SHIPPED")
        except Exception:
            pass

        sh = (
            db.execute(
                select(Shipment)
                .options(selectinload(Shipment.cost_center))
                .where(Shipment.id == sh.id)
            )
            .scalar_one()
        )
        return shipment_to_out(sh, db)

    except Exception:
        db.rollback()
        raise


@app.post("/api/shipments/{shipment_id}/change-shipping", response_model=ShipmentOut)
def change_shipping(
    shipment_id: str,
    body: ShipmentShippingChange,
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(require_reception),
):
    sh = db.get(Shipment, shipment_id)
    if not sh:
        raise HTTPException(404, "Shipment not found")

    if sh.status not in (ShipmentStatus.SHIPPED, ShipmentStatus.SHIPPING_CHANGED):
        raise HTTPException(400, f"Cannot change shipping when status={sh.status}")

    if sh.status in (ShipmentStatus.CANCELLED, ShipmentStatus.CANCELLED_AFTER_SHIPPED):
        raise HTTPException(400, "Cannot change shipping for cancelled shipment")

    carrier_id = (body.carrier_id or "").strip()
    tracking_no = (body.carrier_tracking_no or "").strip()

    if not carrier_id:
        raise HTTPException(400, "carrier_id is required")
    if len(tracking_no) < 2:
        raise HTTPException(400, "carrier_tracking_no is required")

    try:
        old_status = sh.status
        old_carrier_id = sh.carrier_id
        old_tracking_no = sh.carrier_tracking_no

        sh.carrier_id = carrier_id
        sh.carrier_tracking_no = tracking_no
        sh.status = ShipmentStatus.SHIPPING_CHANGED
        sh.updated_at = datetime.utcnow()

        add_event(
            db,
            sh.id,
            "SHIPPING_CHANGED",
            user.upn,
            {
                "old_status": str(old_status).split(".")[-1],
                "new_status": "SHIPPING_CHANGED",
                "old_carrier_id": str(old_carrier_id) if old_carrier_id else None,
                "old_carrier_tracking_no": old_tracking_no,
                "new_carrier_id": carrier_id,
                "new_carrier_tracking_no": tracking_no,
                "shipped_at": str(sh.shipped_at) if sh.shipped_at else None,
            },
        )

        db.commit()
        db.refresh(sh)

        try:
            from .notify import notify_status_change
            sh_mail = (
                db.execute(
                    select(Shipment)
                    .options(selectinload(Shipment.carrier), selectinload(Shipment.cost_center))
                    .where(Shipment.id == sh.id)
                )
                .scalar_one()
            )
            notify_status_change(sh_mail, "SHIPPING_CHANGED")
        except Exception:
            pass

        sh = (
            db.execute(
                select(Shipment)
                .options(selectinload(Shipment.cost_center))
                .where(Shipment.id == sh.id)
            )
            .scalar_one()
        )
        return shipment_to_out(sh, db)

    except Exception:
        db.rollback()
        raise


@app.post("/api/shipments/{shipment_id}/cancel", response_model=ShipmentOut)
def cancel_shipment(
    shipment_id: str,
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(get_current_user),
):
    sh = db.get(Shipment, shipment_id)
    if not sh:
        raise HTTPException(404, "Shipment not found")

    if not user.is_reception and sh.requested_by_upn != user.upn:
        raise HTTPException(403, "Not allowed")

    if sh.status in (ShipmentStatus.CANCELLED, ShipmentStatus.CANCELLED_AFTER_SHIPPED):
        sh = (
            db.execute(
                select(Shipment)
                .options(selectinload(Shipment.cost_center))
                .where(Shipment.id == sh.id)
            )
            .scalar_one()
        )
        return shipment_to_out(sh, db)

    try:
        old_status = sh.status

        if sh.status in (ShipmentStatus.SHIPPED, ShipmentStatus.SHIPPING_CHANGED):
            sh.status = ShipmentStatus.CANCELLED_AFTER_SHIPPED
            event_type = "CANCELLED_AFTER_SHIPPED"
        else:
            sh.status = ShipmentStatus.CANCELLED
            event_type = "CANCELLED"

        sh.updated_at = datetime.utcnow()

        add_event(
            db,
            sh.id,
            event_type,
            user.upn,
            {
                "old_status": str(old_status),
                "new_status": str(sh.status),
                "carrier_id": str(sh.carrier_id) if sh.carrier_id else None,
                "carrier_tracking_no": sh.carrier_tracking_no,
                "shipped_at": str(sh.shipped_at) if sh.shipped_at else None,
            },
        )

        db.commit()
        db.refresh(sh)

        try:
            from .notify import notify_status_change
            sh_mail = (
                db.execute(
                    select(Shipment)
                    .options(selectinload(Shipment.carrier), selectinload(Shipment.cost_center))
                    .where(Shipment.id == sh.id)
                )
                .scalar_one()
            )
            notify_status_change(sh_mail, str(sh.status).split(".")[-1])
        except Exception:
            pass

        sh = (
            db.execute(
                select(Shipment)
                .options(selectinload(Shipment.cost_center))
                .where(Shipment.id == sh.id)
            )
            .scalar_one()
        )
        return shipment_to_out(sh, db)

    except Exception:
        db.rollback()
        raise

# ======================================================
# INCOMING ENDPOINTY (recepcja)
# ======================================================

@app.get("/api/incoming-shipments", response_model=list[IncomingShipmentOut])
def list_incoming_shipments(
    status: ShipmentStatus | None = Query(default=None),
    q: str | None = Query(default=None),
    limit: int = Query(default=200, ge=1, le=500),
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(require_reception),
):
    stmt = (
        select(IncomingShipment)
        .options(selectinload(IncomingShipment.carrier))
        # received_at może być NULL, więc bezpieczniej sortować po COALESCE
        .order_by(func.coalesce(IncomingShipment.received_at, IncomingShipment.created_at).desc())
        .limit(limit)
    )

    if status:
        stmt = stmt.where(IncomingShipment.status == status)

    if q:
        like = f"%{q.strip()}%"
        stmt = stmt.where(
            or_(
                IncomingShipment.internal_no.ilike(like),  # ✅ dodane
                IncomingShipment.carrier_tracking_no.ilike(like),
                IncomingShipment.sender_name.ilike(like),
                IncomingShipment.recipient_name.ilike(like),
                IncomingShipment.recipient_upn.ilike(like),
                IncomingShipment.contents.ilike(like),
            )
        )

    rows = db.execute(stmt).scalars().all()
    return [incoming_to_out(x) for x in rows]


@app.get("/api/incoming-shipments/{incoming_id}", response_model=IncomingShipmentOut)
def get_incoming_shipment(
    incoming_id: str,
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(require_reception),
):
    x = (
        db.execute(
            select(IncomingShipment)
            .options(selectinload(IncomingShipment.carrier))
            .where(IncomingShipment.id == incoming_id)
        )
        .scalar_one_or_none()
    )
    if not x:
        raise HTTPException(404, "Incoming shipment not found")
    return incoming_to_out(x)


@app.post("/api/incoming-shipments", response_model=IncomingShipmentOut)
def register_incoming_shipment(
    body: IncomingShipmentCreate,
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(require_reception),
):
    try:
        internal_no = next_incoming_internal_no(db)

        x = IncomingShipment(
            internal_no=internal_no,
            direction=Direction.INCOMING,
            status=ShipmentStatus.AT_RECEPTION,

            carrier_id=body.carrier_id,
            carrier_tracking_no=body.carrier_tracking_no.strip(),

            sender_name=body.sender_name.strip(),
            recipient_upn=str(body.recipient_upn).strip(),
            recipient_name=body.recipient_name.strip(),
            contents=(body.contents.strip() if body.contents else None),

            received_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
        )

        db.add(x)
        db.commit()
        db.refresh(x)

        # mail: doładuj carrier, żeby miał nazwę
        try:
            from .notify import notify_incoming_registered
            x_mail = (
                db.execute(
                    select(IncomingShipment)
                    .options(selectinload(IncomingShipment.carrier))
                    .where(IncomingShipment.id == x.id)
                )
                .scalar_one()
            )
            notify_incoming_registered(x_mail)
        except Exception:
            pass

        x = (
            db.execute(
                select(IncomingShipment)
                .options(selectinload(IncomingShipment.carrier))
                .where(IncomingShipment.id == x.id)
            )
            .scalar_one()
        )
        return incoming_to_out(x)

    except IntegrityError:
        db.rollback()
        raise HTTPException(409, "Duplicate internal number (retry)")
    except Exception:
        db.rollback()
        raise


@app.post("/api/incoming-shipments/{incoming_id}/picked-up", response_model=IncomingShipmentOut)
def mark_incoming_picked_up(
    incoming_id: str,
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(require_reception),
):
    x = db.get(IncomingShipment, incoming_id)
    if not x:
        raise HTTPException(404, "Incoming shipment not found")

    if x.status == ShipmentStatus.PICKED_UP:
        x = (
            db.execute(
                select(IncomingShipment)
                .options(selectinload(IncomingShipment.carrier))
                .where(IncomingShipment.id == x.id)
            )
            .scalar_one()
        )
        return incoming_to_out(x)

    if x.status == ShipmentStatus.CANCELLED:
        raise HTTPException(400, "Cannot pick up cancelled incoming shipment")

    try:
        x.status = ShipmentStatus.PICKED_UP
        x.picked_up_at = datetime.utcnow()
        x.updated_at = datetime.utcnow()

        db.commit()
        db.refresh(x)

        # mail: doładuj carrier
        try:
            from .notify import notify_incoming_picked_up
            x_mail = (
                db.execute(
                    select(IncomingShipment)
                    .options(selectinload(IncomingShipment.carrier))
                    .where(IncomingShipment.id == x.id)
                )
                .scalar_one()
            )
            notify_incoming_picked_up(x_mail)
        except Exception:
            pass

        x = (
            db.execute(
                select(IncomingShipment)
                .options(selectinload(IncomingShipment.carrier))
                .where(IncomingShipment.id == x.id)
            )
            .scalar_one()
        )
        return incoming_to_out(x)

    except Exception:
        db.rollback()
        raise


# ✅ NOWE: Zmiana odbiorcy tylko dla statusu AT_RECEPTION
@app.post("/api/incoming-shipments/{incoming_id}/change-recipient", response_model=IncomingShipmentOut)
def change_incoming_recipient(
    incoming_id: str,
    body: IncomingRecipientChange,
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(require_reception),
):
    x = db.get(IncomingShipment, incoming_id)
    if not x:
        raise HTTPException(404, "Incoming shipment not found")

    if x.status != ShipmentStatus.AT_RECEPTION:
        raise HTTPException(400, f"Cannot change recipient when status={x.status}")

    upn = (body.recipient_upn or "").strip()
    name = (body.recipient_name or "").strip()

    if len(upn) < 3:
        raise HTTPException(400, "recipient_upn is required")
    if len(name) < 2:
        raise HTTPException(400, "recipient_name is required")

    try:
        old_upn = x.recipient_upn
        old_name = x.recipient_name

        x.recipient_upn = upn
        x.recipient_name = name
        x.updated_at = datetime.utcnow()

        db.commit()
        db.refresh(x)

        # mail: opcjonalnie (jak masz taką funkcję)
        try:
            from .notify import notify_incoming_recipient_changed
            x_mail = (
                db.execute(
                    select(IncomingShipment)
                    .options(selectinload(IncomingShipment.carrier))
                    .where(IncomingShipment.id == x.id)
                )
                .scalar_one()
            )
            notify_incoming_recipient_changed(
                x_mail,
                old_recipient_upn=old_upn,
                old_recipient_name=old_name,
                new_recipient_upn=upn,
                new_recipient_name=name,
                changed_by_upn=user.upn,
            )
        except Exception:
            pass

        x = (
            db.execute(
                select(IncomingShipment)
                .options(selectinload(IncomingShipment.carrier))
                .where(IncomingShipment.id == x.id)
            )
            .scalar_one()
        )
        return incoming_to_out(x)

    except Exception:
        db.rollback()
        raise


# ======================================================
# INCOMING ENDPOINTY (użytkownik)  ✅ NOWE
# ======================================================

@app.get("/api/my-incoming-shipments", response_model=list[IncomingShipmentOut])
def my_incoming_shipments(
    status: ShipmentStatus | None = Query(default=None),
    q: str | None = Query(default=None),
    limit: int = Query(default=200, ge=1, le=500),
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(get_current_user),
):
    """
    Incoming dla zalogowanego użytkownika (po recipient_upn == user.upn).
    """
    stmt = (
        select(IncomingShipment)
        .options(selectinload(IncomingShipment.carrier))
        .where(IncomingShipment.recipient_upn == user.upn)
        .order_by(func.coalesce(IncomingShipment.received_at, IncomingShipment.created_at).desc())
        .limit(limit)
    )

    if status:
        stmt = stmt.where(IncomingShipment.status == status)

    if q:
        like = f"%{q.strip()}%"
        stmt = stmt.where(
            or_(
                IncomingShipment.internal_no.ilike(like),
                IncomingShipment.carrier_tracking_no.ilike(like),
                IncomingShipment.sender_name.ilike(like),
                IncomingShipment.recipient_name.ilike(like),
                IncomingShipment.recipient_upn.ilike(like),
                IncomingShipment.contents.ilike(like),
            )
        )

    rows = db.execute(stmt).scalars().all()
    return [incoming_to_out(x) for x in rows]
